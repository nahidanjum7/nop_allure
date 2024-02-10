import openpyxl

def numRows(file_name,sheet_name):
    Excel_File=openpyxl.load_workbook(file_name) #loading the file
    Sheet=Excel_File[sheet_name] #loading sheet name
    return Sheet.max_row #in this specific we are getting max rows

def readData(file_name,sheet_name,row_no,col_num):
    Excel_File=openpyxl.load_workbook(file_name) #loading the file
    Sheet=Excel_File[sheet_name] #loading sheet name
    return Sheet.cell(row=row_no,column=col_num).value #returning the value of given row and column

def writeData(file_name,sheet_name,row_no,col_no,data):
    Excel_File=openpyxl.load_workbook(file_name) #loading the file
    Sheet=Excel_File[sheet_name] #loading sheet name
    Sheet.cell(row=row_no,column=col_no).value=data #giving the value for given row & column
    Excel_File.save(file_name)

